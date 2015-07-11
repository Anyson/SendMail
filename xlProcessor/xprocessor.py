#coding=utf-8
'''
Created on 2015年7月10日

@author: Anyson
'''

import time
from openpyxl import load_workbook
from xlProcessor.exceptions import InvalidFileException, InvalidExcelException
from xlProcessor.xtag import XTag

class SheetHelper(object):
    '''
    sheet 内容容器
    '''
    def __init__(self):
        self.head = None
        self.rowList = None
    
    def setHead(self, head):
        self.head = head
        self.usernameIndex = -1
        for index in range(0, len(self.head)):
            cell_content = unicode(self.head[index].value)
            if cell_content == u'职员姓名' or cell_content == u'姓名':
                self.usernameIndex = index
                break
    
    def setRowList(self, rowList):
        self.rowList = rowList
    
    def getEmail(self, row):
        return row[0].value
    
    def getUsername(self, row):
        if self.usernameIndex >= 0:
            return row[self.usernameIndex].value
        return ""

class ContentHelper(object):
    ''''
    读取excel后整理合并后的content类
    '''
    def __init__(self, email=None, username=None, content=""):
        self.email = email
        self.username = username
        self.content = content
    
class XProcessor(object):
    '''
    xlsx/xlsm/xltx/xltm 处理器
    '''


    def __init__(self):
        self.content = {}
        
    def loadWorkbook(self, filename=None):
        try:
            self.wb = load_workbook(filename=filename, read_only=True)
            self.sheetnames = self.wb.get_sheet_names()
            self.sheets = {}
            for name in self.sheetnames:
                sheet = self.wb.get_sheet_by_name(name)
                rows = [row for row in sheet.rows]
                if len(rows) > 2 :
                    sheetHelper = SheetHelper()
                    sheet_head = rows[0]
                    sheetHelper.setHead(sheet_head)
                    rows.pop(0)
                    sheetHelper.setRowList(rows)
                    self.sheets[name] = sheetHelper
                    
            if len(self.sheets.keys()) <= 0:
                print u"excel文件的内容为空"
                raise InvalidExcelException(u"excel文件的内容为空")
             
        except Exception:
            err_str = u"打开excel文件%s错误,请检查excel文件是否有错!(支持excel的文件后缀有xlsx/xlsm/xltx/xltm)" % filename
            raise InvalidFileException(err_str)
    
    def loadContentTemplate(self, templateName=None):
        try:
            f = open(templateName, 'rb+')
            self.content_template = f.read()
            f.close()
            self.content_template = unicode(self.content_template, "utf-8")
        except Exception:
            raise InvalidFileException(u"打开模板内容文件%s错误,请检查模板文件是否有错!" % templateName)
        
    def makeContent(self):
        def getTable(heads, contentList):
            tab = u'<table id="mytable" cellspacing="0" >\n\t<tr>\n'
            for th in heads[1:len(heads)]:
                tab += u'\t\t<th scope="col">' + unicode(th.value) + u'</th>\n'
            tab += u"\t</tr>\n\t<tr>\n"
            for td in contentList[1:len(contentList)]:
                tab += u'\t\t<td class="row">' + unicode(td.value) + u'</td>\n'
            tab += u'\t</tr>\n</table> '
            return unicode(tab)
        
        current_time = time.localtime(time.time())
        current_time = time.strftime(u'%Y-%m-%d %H:%M:%S', current_time)
        for sheetname in self.sheets.keys():
            sheetHelper = self.sheets.get(sheetname)
            l = []
            for row in sheetHelper.rowList:
                email = sheetHelper.getEmail(row)
                if not email:
                    continue
                email = unicode(email)
                username = unicode(sheetHelper.getUsername(row))
                content = XTag.parse(self.content_template, { u"username" : username,
                                                              u"tablecontent" : getTable(sheetHelper.head, row),
                                                              u"currenttime" : unicode(current_time)})
                
                contentHelper = ContentHelper(email, username, content)
                l.append(contentHelper)
            self.content[sheetname] = l
            
        
        
    def start(self, filepath, contentTemplate="content.txt"):
        try:
            self.loadWorkbook(filepath)
            self.loadContentTemplate(contentTemplate)
            self.makeContent()
        except Exception, e:
            raise e
            

if __name__ == "__main__":
    xl = XProcessor()
    xl.loadWorkbook("../sendmail/salary.xlsx")
    xl.loadContentTemplate("../sendmail/content.txt")
    xl.makeContent()
    